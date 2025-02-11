import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

# Load the dataset
file_path = "MTA_Ridership_Data_02-07-2025_23-25-21.csv"  # Update to your actual file path
df = pd.read_csv(file_path)

# Convert transit_timestamp to datetime format
df["transit_timestamp"] = pd.to_datetime(df["transit_timestamp"], errors="coerce")

# Extract year, month, and hour
df["year"] = df["transit_timestamp"].dt.year
df["month"] = df["transit_timestamp"].dt.month
df["hour"] = df["transit_timestamp"].dt.hour

# Convert military time (24-hour format) to 12-hour AM/PM format
df["AM_PM"] = df["hour"].apply(lambda x: f"{x % 12 if x % 12 != 0 else 12} {'AM' if x < 12 else 'PM'}")

# Process data for both 2023 and 2024
for year in [2023, 2024]:
    df_year = df[df["year"] == year]

    # Get the unique stations and months
    stations = df_year["station_complex_id"].unique()
    months = sorted(df_year["month"].unique())

    # Store file paths for user download
    excel_files = []

    # Loop through each month and create separate Excel sheets
    for month in months:
        # Filter data for the current month
        month_data = df_year[df_year["month"] == month]

        # Calculate average ridership per station per month
        avg_ridership = month_data.groupby(["station_complex_id", "station_complex", "hour"])["ridership"].mean().reset_index()

        # Create an Excel workbook for this month
        wb = Workbook()
        del wb[wb.sheetnames[0]]  # Remove the default sheet

        # Plot each station's data for the current month
        for station_id in stations:
            station_data = avg_ridership[avg_ridership["station_complex_id"] == station_id]

            if station_data.empty:
                continue

            station_name = station_data["station_complex"].iloc[0]

            # Sanitize station names for Excel sheet titles
            sanitized_station_name = station_name.replace("/", "-").replace("\\", "-").replace("*", "-").replace("[", "(").replace("]", ")").replace(":", "-").replace("?", "-").replace("'", "-").replace(",", "-").replace(".", "-").strip()[:31]

            # Create a line chart
            plt.figure(figsize=(10, 6))
            plt.plot(station_data["hour"], station_data["ridership"], marker='o', linestyle='-', label=f"Avg Riders: {station_data['ridership'].mean():.2f}")
            plt.title(f"Avg Ridership for {sanitized_station_name} - {month}/{year}", fontsize=16)
            plt.xlabel("Hour of Day", fontsize=14)
            plt.ylabel("Avg Riders", fontsize=14)
            plt.grid(True, linestyle='--', alpha=0.6)
            plt.xticks(rotation=45)
            plt.legend()
            plt.tight_layout()

            # Save the plot as an image
            img_name = f"station_{station_id}_month_{month}_{year}.png"
            plt.savefig(img_name)
            plt.close()

            # Create a new sheet in Excel for the station
            ws = wb.create_sheet(title=sanitized_station_name)  # Excel sheet names are limited to 31 characters
            ws.append([f"Station: {station_name} (ID: {station_id}) - {month}/{year}"])
            ws.append(["Hour", "Average Ridership"])

            # Add average ridership data to Excel
            for index, row in station_data.iterrows():
                ws.append([row["hour"], row["ridership"]])

            # Embed the image in the station's tab
            img = Image(img_name)
            ws.add_image(img, "A5")

        # Save the workbook for the current month
        excel_path = f"station_charts_month_{month}_{year}.xlsx"
        wb.save(excel_path)
        excel_files.append(excel_path)

        # Clean up image files
        for station_id in stations:
            img_file = f"station_{station_id}_month_{month}_{year}.png"
            if os.path.exists(img_file):
                os.remove(img_file)

    print(f"Excel files generated for {year}:", excel_files)
