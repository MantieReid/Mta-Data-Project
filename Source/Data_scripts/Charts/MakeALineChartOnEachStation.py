import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import os

# Makes a line chart for each station and embeds it in an Excel workbook, with each station on a separate tab. The data is in a CSV file with columns: station_complex_id, station_complex, hour, ridership.
# Does not save it in EST time. That is in  a separate script.


# Load the data
file_path = "avg_ridership_per_hour.csv"  # Update to the path of your dataset
df = pd.read_csv(file_path)

# Analyze the structure of the data
# Columns: station_complex_id, station_complex, hour, ridership

# Group by station_complex_id
stations = df["station_complex_id"].unique()

# Create an Excel workbook
wb = Workbook()
del wb[wb.sheetnames[0]]  # Remove the default sheet

# Plot each station's data
for station_id in stations:
    station_data = df[df["station_complex_id"] == station_id]
    station_name = station_data["station_complex"].iloc[0]

    # Replace invalid characters in station name
    sanitized_station_name = station_name.replace("/", "-").replace("\\", "-").replace("*", "-").replace("[", "(").replace("]", ")").replace(":", "-").replace("?", "-").replace("'", "-").replace(",", "-").replace(".", "-").strip()[:31]

    # Create a line chart for the station
    plt.figure(figsize=(10, 6))
    plt.plot(station_data["hour"], station_data["ridership"], marker='o', linestyle='-', label=f"Station ID {station_id}")
    plt.title(f"Ridership for {sanitized_station_name}", fontsize=16)
    plt.xlabel("Hour of Day", fontsize=14)
    plt.ylabel("Riders", fontsize=14)
    plt.grid(True, linestyle='--', alpha=0.6)
    plt.xticks(range(0, 24))
    plt.legend()
    plt.tight_layout()

    # Save the plot as an image
    img_name = f"station_{station_id}.png"
    plt.savefig(img_name)
    plt.close()

    # Create a new sheet in Excel for the station
    ws = wb.create_sheet(title=sanitized_station_name)  # Excel sheet names are limited to 31 characters
    ws.append([f"Station: {station_name} (ID: {station_id})"])

    # Embed the image in the station's tab
    img = Image(img_name)
    ws.add_image(img, "A2")

# Save the workbook
excel_path = "station_charts_by_tab.xlsx"
wb.save(excel_path)

# Clean up image files
for station_id in stations:
    os.remove(f"station_{station_id}.png")

print(f"Charts have been embedded in {excel_path}, with each station on a separate tab.")
