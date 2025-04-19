import pandas as pd
import os
import gc
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


def process_data_in_chunks(file_path, years_to_analyze, chunk_size=100000):
    """
    Process the MTA ridership data in chunks to reduce memory usage
    
    Args:
        file_path (str): Path to the CSV file
        years_to_analyze (list): List of years to extract
        chunk_size (int): Number of rows to process at once
        
    Returns:
        dict: Dictionary with years as keys and processed dataframes as values
    """
    # Initialize dictionaries to store results
    year_sums = {year: {} for year in years_to_analyze}
    year_counts = {year: {} for year in years_to_analyze}
    
    # Define the columns we need - this reduces memory by loading only necessary columns
    usecols = ['transit_timestamp', 'station_complex_id', 'station_complex', 
               'ridership']
    
    print(f"Loading and processing data from {file_path} in chunks...")
    
    # Read a sample to inspect the data
    sample = pd.read_csv(file_path, nrows=5)
    print("Sample data:")
    print(sample[usecols].head() if all(col in sample.columns for col in usecols) else "Missing required columns")
    
    # Check if required columns exist
    missing_cols = [col for col in usecols if col not in sample.columns]
    if missing_cols:
        print(f"Error: Missing required columns: {missing_cols}")
        print(f"Available columns: {sample.columns.tolist()}")
        return {}
    
    # Process the file in chunks
    rows_processed = 0
    rows_matching_years = 0
    
    for chunk_num, chunk in enumerate(pd.read_csv(file_path, chunksize=chunk_size, usecols=usecols)):
        if chunk_num % 10 == 0:
            print(f"Processing chunk {chunk_num}...")
        
        # Clean data
        # Ensure ridership is numeric
        chunk['ridership'] = pd.to_numeric(chunk['ridership'], errors='coerce')
        
        # Convert timestamp to datetime 
        chunk['transit_timestamp'] = pd.to_datetime(chunk['transit_timestamp'], errors='coerce')
        
        # Drop rows with invalid timestamps or ridership
        chunk = chunk.dropna(subset=['transit_timestamp', 'ridership'])
        
        # Extract year and hour from timestamp
        chunk['year'] = chunk['transit_timestamp'].dt.year
        chunk['hour'] = chunk['transit_timestamp'].dt.hour
        
        rows_processed += len(chunk)
        
        # Filter for relevant years
        for year in years_to_analyze:
            year_chunk = chunk[chunk['year'] == year]
            
            if chunk_num % 10 == 0 and len(year_chunk) > 0:
                print(f"Chunk {chunk_num}: Found {len(year_chunk)} rows for year {year}")
            
            rows_matching_years += len(year_chunk)
            
            if not year_chunk.empty:
                # Group by required columns and calculate sum and count
                grouped = year_chunk.groupby([
                    'station_complex_id', 
                    'station_complex', 
                    'hour'
                ])['ridership'].agg(['sum', 'count']).reset_index()
                
                # Process each group
                for _, row in grouped.iterrows():
                    key = (row['station_complex_id'], row['station_complex'], row['hour'])
                    
                    if key in year_sums[year]:
                        year_sums[year][key] += row['sum']
                        year_counts[year][key] += row['count']
                    else:
                        year_sums[year][key] = row['sum']
                        year_counts[year][key] = row['count']
        
        # Explicitly delete chunk to free memory
        del chunk
        gc.collect()
    
    print(f"\nProcessed {rows_processed} total rows")
    print(f"Found {rows_matching_years} rows matching target years {years_to_analyze}")
    
    # Calculate averages and create final dataframes
    result = {}
    for year in years_to_analyze:
        rows = []
        key_count = len(year_sums[year])
        print(f"Creating dataframe for year {year} with {key_count} station/hour combinations")
        
        for key, total in year_sums[year].items():
            count = year_counts[year][key]
            avg = total / count if count > 0 else 0
            station_id, station_name, hour = key
            rows.append({
                'station_complex_id': station_id,
                'station_complex': station_name,
                'hour': int(hour),
                'ridership': float(avg)
            })
        
        if rows:
            result[year] = pd.DataFrame(rows)
            # Sort by station_complex and hour for better readability
            result[year] = result[year].sort_values(by=['station_complex', 'hour'])
            print(f"Created dataframe for year {year} with {len(rows)} rows")
        else:
            print(f"Warning: No data found for year {year}")
            result[year] = pd.DataFrame(columns=['station_complex_id', 'station_complex', 'hour', 'ridership'])
    
    return result


def save_results_to_excel(data_by_year, prefix="avg_ridership"):
    """
    Save dataframes to Excel files with table formatting (Dark Teal, Table Style Medium 2)
    
    Args:
        data_by_year (dict): Dictionary with years as keys and dataframes as values
        prefix (str): Prefix for output filenames
        
    Returns:
        list: List of output filenames
    """
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    file_path_output = os.path.join(base_dir, "Data", "reports")
    
    if not os.path.exists(file_path_output):
        os.makedirs(file_path_output)
    
    current_time = datetime.now()
    date_time_str = current_time.strftime("%B %d, %Y %I-%M %p")
    
    filenames = []
    years = list(data_by_year.keys())
    
    for year in years:
        # Make a copy of the dataframe to avoid modifying the original
        df = data_by_year[year].copy()
        
        # Check for duplicate station-hour combinations
        duplicates = df.duplicated(subset=['station_complex', 'hour'], keep=False)
        if duplicates.any():
            print(f"Warning: Found {duplicates.sum()} duplicate station-hour combinations for year {year}")
            print("Aggregating duplicate entries...")
            
            # Aggregate the duplicates by taking the mean of ridership values
            df = df.groupby(['station_complex', 'hour']).agg({
                'station_complex_id': 'first',  # Take the first ID
                'ridership': 'mean'  # Average the ridership values
            }).reset_index()

        # Double-check there are no duplicates after aggregation
        if df.duplicated(subset=['station_complex', 'hour']).any():
            print("ERROR: Still found duplicates after aggregation!")
        else:
            print(f"Successfully removed duplicates. DataFrame now has {len(df)} rows.")
        
        # Only keep columns needed for Excel output
        excel_df = df[['station_complex', 'hour', 'ridership']].copy()
        
        # Rename columns for better readability in the Excel table
        excel_df = excel_df.rename(columns={
            'station_complex': 'Station Name',
            'hour': 'Hour',
            'ridership': 'Average Ridership'
        })
        
        # Create Excel file path with local date and time included
        filename = os.path.join(file_path_output, f"{prefix}_{year}Made_On_{date_time_str}.xlsx")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Ridership {year}"
        
        # Write headers
        headers = list(excel_df.columns)
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        # Write data
        for row_idx, row in enumerate(dataframe_to_rows(excel_df, index=False, header=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                # Format the "Average Ridership" column to display as number with 2 decimal places
                if col_idx == 3:  # The "Average Ridership" column
                    cell.number_format = '0.00'
        
        # Create a table
        table_ref = f"A1:{get_column_letter(len(headers))}{len(excel_df) + 1}"
        table = Table(displayName=f"RidershipTable{year}", ref=table_ref)
        
        # Set table style to "Table Style Medium 2" (Dark Teal)
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        
        # Add the table to the worksheet
        ws.add_table(table)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(filename)
        print(f"Saved Excel file with formatted table: {filename}")
        
        filenames.append(filename)
    
    return filenames, file_path_output


def create_json_export(data_by_year, file_path_output):
    """
    Create JSON export from the already processed data
    
    Args:
        data_by_year (dict): Dictionary with years as keys and dataframes as values
        file_path_output (str): Path to save the JSON file
        
    Returns:
        str: Path to the saved JSON file
    """
    import json
    import numpy as np
    
    # Create a dictionary to store JSON data
    json_output = {}
    
    # Get current date and time for filename
    current_datetime = datetime.now().strftime("%Y_%m_%d_%I_%M_%S_%p")
    
    # Process each year's data
    for year, df in data_by_year.items():
        # Create a copy to avoid modifying the original
        year_df = df.copy()
        
        # Check for duplicate station-hour combinations
        duplicates = year_df.duplicated(subset=['station_complex', 'hour'], keep=False)
        if duplicates.any():
            print(f"Warning: Found {duplicates.sum()} duplicate station-hour combinations for year {year} in JSON export")
            print("Aggregating duplicate entries...")
            
            # Aggregate the duplicates by taking the mean of ridership values
            year_df = year_df.groupby(['station_complex', 'hour']).agg({
                'station_complex_id': 'first',  # Take the first ID
                'ridership': 'mean'  # Average the ridership values
            }).reset_index()
            
            print(f"After aggregation, dataframe has {len(year_df)} rows")
        
        # Manually create a wide-format dataframe instead of using pivot
        # This avoids potential issues with pandas pivot function
        
        # First, get unique stations
        unique_stations = year_df['station_complex'].unique()
        
        # Create a dictionary to hold the transformed data
        transformed_data = []
        
        for station in unique_stations:
            # Filter data for this station
            station_data = year_df[year_df['station_complex'] == station]
            
            # Create a record for this station
            station_record = {'station_complex': station}
            
            # Add hourly data
            for _, row in station_data.iterrows():
                hour = row['hour']
                ridership = row['ridership']
                station_record[f'hour_{hour}'] = ridership
            
            # Calculate the daily average
            station_record['daily_average'] = station_data['ridership'].mean()
            
            # Add to the transformed data
            transformed_data.append(station_record)
        
        # Add to JSON output
        json_output[f"{year} Average Ridership"] = transformed_data
    
    # Handle numpy data types for JSON serialization
    def convert_nan(obj):
        if isinstance(obj, dict):
            return {k: convert_nan(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [convert_nan(item) for item in obj]
        elif isinstance(obj, np.ndarray):
            return convert_nan(obj.tolist())
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif pd.isna(obj):
            return None
        else:
            return obj

    json_output = convert_nan(json_output)

    # Generate a filename with timestamp
    json_filename = f"MTA_Hourly_Average_Ridership_{current_datetime}.json"
    json_file_path = os.path.join(file_path_output, json_filename)

    # Save JSON file
    with open(json_file_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_output, json_file, indent=2)

    print(f"JSON data saved to: {json_file_path}")
    
    return json_file_path


def main():
    """
    Main function to execute the MTA ridership analysis pipeline
    """
    # File path
    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    file_path = os.path.join(base_dir, "Data", "Raw", "MTA_Subway_Hourly_Ridership__2020-2024.csv")
    
    # Check if file exists
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return
    
    # Years to analyze
    years_to_analyze = [2023, 2024]
    
    print(f"Looking for data from years: {years_to_analyze}")
    
    # Process data in chunks and calculate average ridership - DO THIS ONLY ONCE
    avg_ridership_by_year = process_data_in_chunks(file_path, years_to_analyze)
    
    if not avg_ridership_by_year:
        print("Error: No data processed, check previous messages for details")
        return
    
    # Save results to Excel with table formatting
    print("Saving results to Excel with table formatting...")
    saved_files, file_path_output = save_results_to_excel(avg_ridership_by_year)
    
    # Create JSON export from the SAME processed data
    print("Creating JSON export...")
    json_file_path = create_json_export(avg_ridership_by_year, file_path_output)
    
    # Display the results
    for filename in saved_files:
        print(f"Average ridership saved to {filename}")
    
    print(f"JSON data saved to: {json_file_path}")
    print("Processing complete!")


if __name__ == "__main__":
    main()